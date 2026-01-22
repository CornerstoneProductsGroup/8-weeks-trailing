
def _file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------- AUTOLOAD_VENDOR_MAP ----------
# Load Vendor-SKU pricing map automatically from local file on startup
def autoload_vendor_map(conn):
    import pandas as pd, os
    map_path = os.path.join(os.path.dirname(__file__), "Vendor-SKU Map.xlsx")
    if not os.path.exists(map_path):
        return
    df = pd.read_excel(map_path)
    df.columns = [c.strip().lower() for c in df.columns]
    df = df.rename(columns={
        "retailer": "retailer",
        "vendor": "vendor",
        "sku": "sku",
        "price": "unit_price",
        "unit price": "unit_price",
    })
    df = df[["retailer","vendor","sku","unit_price"]].dropna()
    df["retailer"] = df["retailer"].astype(str).str.strip()
    df["vendor"] = df["vendor"].astype(str).str.strip()
    df["sku"] = df["sku"].astype(str).str.strip()
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0)

    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sku_mapping (
            retailer TEXT,
            vendor TEXT,
            sku TEXT,
            unit_price REAL,
            UNIQUE(retailer, vendor, sku)
        )
    """)
    for _,r in df.iterrows():
        cur.execute(
            "INSERT OR REPLACE INTO sku_mapping (retailer,vendor,sku,unit_price) VALUES (?,?,?,?)",
            (r["retailer"], r["vendor"], r["sku"], float(r["unit_price"]))
        )
    conn.commit()
# ---------- END AUTOLOAD_VENDOR_MAP ----------

COMPACT_TABLE_CSS = """
<style>
/* Reduce padding inside Streamlit dataframes */
div[data-testid="stDataFrame"] table td,
div[data-testid="stDataFrame"] table th {
  padding: 0.25rem 0.4rem !important;
  white-space: nowrap;
}

/* Prevent tables from stretching wide */
div[data-testid="stDataFrame"] {
  max-width: 600px;
}
</style>
"""

SPLIT_TABLE_CSS = """
<style>
/* Slightly reduce spacing between Streamlit columns */
div[data-testid="stHorizontalBlock"] { gap: 0.5rem !important; }
</style>
"""
def init_meta(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_meta (
            k TEXT PRIMARY KEY,
            v TEXT
        )
        """
    )
    conn.commit()

def get_meta(conn, k: str):
    init_meta(conn)
    row = conn.execute("SELECT v FROM app_meta WHERE k = ?", (k,)).fetchone()
    return row[0] if row else None

def set_meta(conn, k: str, v: str):
    init_meta(conn)
    conn.execute("INSERT INTO app_meta(k, v) VALUES(?, ?) ON CONFLICT(k) DO UPDATE SET v=excluded.v", (k, v))
    conn.commit()

def ensure_mapping_loaded(conn, mapping_path: str):
    """
    Ensures sku_mapping is populated from the bundled vendor map Excel.
    Reloads automatically if the file changes.
    """
    init_meta(conn)
    file_hash = _file_md5(mapping_path) if os.path.exists(mapping_path) else None
    prev_hash = get_meta(conn, "mapping_hash")

    # If table empty OR file changed, import
    try:
        cnt = conn.execute("SELECT COUNT(*) FROM sku_mapping WHERE active = 1").fetchone()[0]
    except Exception:
        cnt = 0

    if (cnt == 0) or (file_hash and prev_hash != file_hash):
        # Re-import mapping (uses existing import function if present)
        df_map = pd.read_excel(mapping_path)
        # Normalize expected columns
        df_map.columns = [str(c).strip() for c in df_map.columns]
        # Try common names
        col_retailer = next((c for c in df_map.columns if c.lower() in ["retailer", "store"]), None)
        col_vendor = next((c for c in df_map.columns if c.lower() in ["vendor", "manufacturer"]), None)
        col_sku = next((c for c in df_map.columns if c.lower() in ["sku", "vendor sku", "retailer sku"]), None)
        col_price = next((c for c in df_map.columns if "price" in c.lower()), None)

        if not (col_retailer and col_vendor and col_sku):
            raise ValueError("Vendor map is missing required columns (Retailer, Vendor, SKU).")

        df_norm = pd.DataFrame({
            "retailer": df_map[col_retailer].astype(str).str.strip(),
            "vendor": df_map[col_vendor].astype(str).str.strip(),
            "sku": df_map[col_sku].astype(str).str.strip(),
        })
        if col_price:
            df_norm["unit_price"] = pd.to_numeric(df_map[col_price], errors="coerce")
        else:
            df_norm["unit_price"] = pd.NA

        df_norm = df_norm[df_norm["sku"].ne("") & df_norm["retailer"].ne("")].copy()
        df_norm["active"] = 1
        df_norm["sort_order"] = pd.NA

        # De-duplicate to avoid UNIQUE/PK conflicts (keep first)
        df_norm = df_norm.drop_duplicates(subset=["retailer", "sku"], keep="first")

        # Replace existing mapping table atomically
        df_norm.to_sql("sku_mapping", conn, if_exists="replace", index=False)
        conn.commit()
        conn.execute("CREATE INDEX IF NOT EXISTS idx_sku_mapping_retailer_sku ON sku_mapping(retailer, sku)")
        conn.commit()
        if file_hash:
            set_meta(conn, "mapping_hash", file_hash)
import streamlit as st
import pandas as pd
import hashlib
import sqlite3
import json
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
import os
import re

def parse_week_label_from_filename(filename: str):
    """
    Accepts filenames like:
      'APP 1-1 thru 1-2.xlsx'
      'APP 1-5 thru 1-9.xlsx'
    Returns (label, start_date, end_date) for 2026, or (None,None,None) if not matched.
    """
    base = os.path.basename(filename)
    m = re.search(r'APP\s+(\d{1,2})-(\d{1,2})\s+thru\s+(\d{1,2})-(\d{1,2})', base, re.IGNORECASE)
    if not m:
        return None, None, None
    m1, d1, m2, d2 = map(int, m.groups())
    try:
        start = date(2026, m1, d1)
        end = date(2026, m2, d2)
    except Exception:
        return None, None, None
    label = f"{m1}-{d1} / {m2}-{d2}"
    return label, start, end

def norm_name(s: str):
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())

def normalize_retailer_sheet_name(sheet_name: str):
    """
    Maps uploaded sheet names to your retailer names used in the mapping.
    Adjust here if your mapping uses different naming.
    """
    n = norm_name(sheet_name)
    if n in ("depot", "homedepot", "thehomedepot"):
        return "Depot"
    if n in ("lowes", "lowesinc"):
        return "Lowe's"
    if n == "amazon":
        return "Amazon"
    if n in ("tractorsupply", "tsc", "tractorsupplyco"):
        return "Tractor Supply"
    if n in ("depotso", "depotspecialorders", "specialorders"):
        return "Depot SO"
    return sheet_name  # fallback

def parse_app_aggregated_sheet(ws):
    """
    APP workbook format: SKU in col A, Units in col B (already aggregated).
    Skips title rows like 'Depot '.
    """
    out = {}
    for r in range(1, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        qty = ws.cell(r, 2).value
        if sku is None:
            continue
        sku_str = str(sku).strip()
        if sku_str == "" or sku_str.lower() in ("sku", "vendor style"):
            continue
        # skip title row like "Depot "
        if qty is None and len(sku_str) <= 20 and sku_str.lower().strip() in ("depot", "depot ", "lowe's", "amazon", "tractor supply", "depot so"):
            continue
        try:
            q = float(qty)
        except Exception:
            continue
        # keep zeros too? We'll ignore zeros to avoid clutter
        if q == 0:
            continue
        out[sku_str] = out.get(sku_str, 0.0) + q
    return out
from pathlib import Path

APP_TITLE = "Weekly Retailer Report (Multi-week View)"
DB_PATH = "app.db"
APP_DIR = Path(__file__).resolve().parent

# -----------------------------
# Week selector (2026 only for now)
# -----------------------------
def weeks_2026():
    rows = []
    # Special partial week
    rows.append((date(2026, 1, 1), date(2026, 1, 2), "1-1 / 1-2"))
    monday = date(2026, 1, 5)
    for i in range(0, 60):
        start = monday + timedelta(weeks=i)
        end = start + timedelta(days=4)
        if start.year != 2026:
            break
        if end.year != 2026:
            end = date(2026, 12, 31)
        rows.append((start, end, f"{start.month}-{start.day} / {end.month}-{end.day}"))
        if end == date(2026, 12, 31):
            break
    return rows


MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def weeks_for_months(week_meta, months_selected):
    if not months_selected:
        return []
    labels = []
    for start, _, lbl in week_meta:
        if start.month in months_selected:
            labels.append(lbl)
    return labels


def fmt_currency_str(x):
    """Format a number like $3,234.00 and negatives like ($1,174.95)."""
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        v = float(x)
    except Exception:
        return ""
    s = f"${abs(v):,.2f}"
    return f"({s})" if v < 0 else s

# -----------------------------
# DB helpers
# -----------------------------
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def init_db(conn: sqlite3.Connection):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS sku_mapping (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        retailer TEXT NOT NULL,
        vendor TEXT NOT NULL,
        sku TEXT NOT NULL,
        unit_price REAL,
        active INTEGER NOT NULL DEFAULT 1,
        sort_order INTEGER,
        UNIQUE(retailer, sku)
    );

    CREATE TABLE IF NOT EXISTS weekly_results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        week_start TEXT NOT NULL,
        week_end TEXT NOT NULL,
        retailer TEXT NOT NULL,
        sku TEXT NOT NULL,
        units_auto REAL,
        units_override REAL,
        sales_manual REAL,
        notes TEXT,
        updated_at TEXT NOT NULL,
        UNIQUE(week_start, retailer, sku)
    );

    CREATE INDEX IF NOT EXISTS idx_weekly_results_week_retailer
    ON weekly_results(week_start, retailer);

    CREATE TABLE IF NOT EXISTS ui_state (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    """)
    # In case an older DB exists, try to add unit_price (no-op if already present)
    try:
        conn.execute("ALTER TABLE sku_mapping ADD COLUMN unit_price REAL;")
    except Exception:
        pass
    conn.commit()

def mapping_count(conn):
    df = pd.read_sql_query("SELECT COUNT(*) AS n FROM sku_mapping", conn)
    return int(df.loc[0, "n"]) if not df.empty else 0

def get_ui_state(conn, key: str, default=None):
    try:
        df = pd.read_sql_query("SELECT value FROM ui_state WHERE key = ?", conn, params=(key,))
        if df.empty:
            return default
        return json.loads(df.loc[0, "value"])
    except Exception:
        return default

def set_ui_state(conn, key: str, value):
    try:
        conn.execute(
            "INSERT INTO ui_state(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, json.dumps(value))
        )
        conn.commit()
    except Exception:
        pass

def mapping_has_any_price(conn) -> bool:
    try:
        df = pd.read_sql_query("SELECT COUNT(*) AS n FROM sku_mapping WHERE unit_price IS NOT NULL", conn)
        return int(df.loc[0, "n"]) > 0
    except Exception:
        return False

def refresh_mapping_from_bundled_if_needed(conn):
    """
    If mapping exists but has no prices populated, reload from bundled Vendor-SKU Map.xlsx.
    This fixes the common case where the DB was bootstrapped from an older map without price.
    """
    if mapping_count(conn) == 0:
        return False
    if mapping_has_any_price(conn):
        return False
    # Try reading bundled map relative to app.py
    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        Path("Vendor-SKU Map.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                # only reload if the file actually contains a price column
                price_col = next((c for c in df_map.columns if "price" in str(c).lower()), None)
                if price_col:
                    upsert_mapping(conn, df_map)
                    return True
        except Exception:
            continue
    return False

def upsert_mapping(conn, df: pd.DataFrame):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Retailer", "SKU", "Vendor"}
    if not required.issubset(set(df.columns)):
        raise ValueError(f"Mapping must contain columns: {sorted(required)}. Found: {list(df.columns)}")

    price_col = next((c for c in df.columns if "price" in str(c).lower()), None)

    df = df[list(required) + ([price_col] if price_col else [])].dropna(subset=["Retailer", "SKU", "Vendor"])
    df["Retailer"] = df["Retailer"].astype(str).str.strip()
    df["SKU"] = df["SKU"].astype(str).str.strip()
    df["Vendor"] = df["Vendor"].astype(str).str.strip()

    if price_col:
        # coerce to numeric
        df[price_col] = pd.to_numeric(df[price_col], errors="coerce")

    retailers = sorted(df["Retailer"].unique().tolist())
    cur = conn.cursor()
    cur.executemany("DELETE FROM sku_mapping WHERE retailer = ?", [(r,) for r in retailers])

    rows = []
    for r in retailers:
        sub = df[df["Retailer"] == r].reset_index(drop=True)
        for i, row in sub.iterrows():
            price = float(row[price_col]) if price_col and pd.notna(row[price_col]) else None
            rows.append((row["Retailer"], row["Vendor"], row["SKU"], price, 1, i + 1))

    cur.executemany("""
        INSERT INTO sku_mapping(retailer, vendor, sku, unit_price, active, sort_order)
        VALUES(?,?,?,?,?,?)
    """, rows)
    conn.commit()

def bootstrap_default_mapping(conn):
    """
    Load a bundled mapping file shipped with the app on first run (or when DB is empty).
    Works on Streamlit Cloud where the working directory can vary.
    """
    if mapping_count(conn) > 0:
        return False

    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        APP_DIR / "Vendor-SKU Map - example.xlsx",
        Path("Vendor-SKU Map.xlsx"),
        Path("Vendor-SKU Map - example.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                upsert_mapping(conn, df_map)
                return True
        except Exception:
            continue
    return False

    candidates = [
        APP_DIR / "Vendor-SKU Map.xlsx",
        APP_DIR / "Vendor-SKU Map - example.xlsx",
        Path("Vendor-SKU Map.xlsx"),
        Path("Vendor-SKU Map - example.xlsx"),
    ]
    for p in candidates:
        try:
            if p.exists():
                df_map = pd.read_excel(p, sheet_name=0)
                upsert_mapping(conn, df_map)
                return True
        except Exception:
            continue
    return False
    for fn in ["Vendor-SKU Map.xlsx", "Vendor-SKU Map - example.xlsx"]:
        try:
            df_map = pd.read_excel(fn, sheet_name=0)
            upsert_mapping(conn, df_map)
            return True
        except Exception:
            continue
    return False

def get_retailers(conn):
    df = pd.read_sql_query("""
        SELECT DISTINCT retailer FROM sku_mapping
        WHERE active = 1
        ORDER BY retailer
    """, conn)
    return df["retailer"].tolist()

def get_mapping_for_retailer(conn, retailer: str):
    return pd.read_sql_query("""
        SELECT vendor, sku, unit_price, sort_order
        FROM sku_mapping
        WHERE active = 1 AND retailer = ?
        ORDER BY COALESCE(sort_order, 999999), vendor, sku
    """, conn, params=(retailer,))

def get_week_records(conn, retailer: str, week_starts: list[str]):
    if not week_starts:
        return pd.DataFrame(columns=["week_start","sku","units_auto","units_override","sales_manual","notes"])
    placeholders = ",".join(["?"] * len(week_starts))
    q = f"""
        SELECT week_start, sku, units_auto, units_override, sales_manual, notes
        FROM weekly_results
        WHERE retailer = ? AND week_start IN ({placeholders})
    """
    return pd.read_sql_query(q, conn, params=[retailer] + week_starts)

def set_units_auto_from_upload(conn, week_start: date, week_end: date, retailer: str, units_by_sku: dict):
    now = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    cur = conn.cursor()
    for sku, units in units_by_sku.items():
        sku = str(sku).strip()
        try:
            units_val = float(units)
        except Exception:
            continue
        cur.execute("""
            INSERT INTO weekly_results(week_start, week_end, retailer, sku,
                                      units_auto, units_override, sales_manual, notes, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(week_start, retailer, sku) DO UPDATE SET
                week_end=excluded.week_end,
                units_auto=excluded.units_auto,
                updated_at=excluded.updated_at
        """, (
            week_start.isoformat(), week_end.isoformat(), retailer, sku,
            units_val, None, None, None, now
        ))
    conn.commit()

# -----------------------------
# Upload parser (based on your example workbook)
# -----------------------------
def parse_weekly_workbook(file, sheet_name: str):
    wb = load_workbook(file, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out = {}

    def add(sku, qty):
        if sku is None:
            return
        sku = str(sku).strip()
        if sku == "" or sku.lower() == "sku":
            return
        try:
            q = float(qty)
        except Exception:
            return
        out[sku] = out.get(sku, 0.0) + q

    if sheet_name in ("Depot", "Lowe's"):
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 5).value, ws.cell(r, 6).value)  # E, F

    elif sheet_name == "Depot SO":
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 4).value, ws.cell(r, 5).value)  # D, E

    elif sheet_name == "Amazon":
        for r in range(1, ws.max_row + 1):
            add(ws.cell(r, 3).value, ws.cell(r, 14).value)  # C, N

    elif sheet_name == "TSC":
        header_row = None
        for r in range(1, min(ws.max_row, 10) + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if isinstance(a, str) and isinstance(b, str) and "Vendor" in a and "Qty" in b:
                header_row = r
                break
        start = (header_row + 1) if header_row else 2
        for r in range(start, ws.max_row + 1):
            add(ws.cell(r, 1).value, ws.cell(r, 2).value)

    return {k: v for k, v in out.items() if v != 0}

# -----------------------------
# Build multi-week view dataframe
# -----------------------------
def build_multiweek_df(conn, retailer: str, week_meta: list[tuple[date,date,str]], display_labels: list[str], edit_label: str):
    mapping = get_mapping_for_retailer(conn, retailer)
    if mapping.empty:
        return pd.DataFrame()

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    starts = [label_to_start[lbl] for lbl in display_labels if lbl in label_to_start]
    wk = get_week_records(conn, retailer, starts)

    # resolved units per (week_start, sku)
    if not wk.empty:
        wk["UnitsResolved"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
    else:
        wk = pd.DataFrame(columns=["week_start","sku","UnitsResolved","sales_manual","notes"])

    base = mapping.rename(columns={"vendor":"Vendor","sku":"SKU","unit_price":"Unit Price"}).copy()
    base["Unit Price"] = pd.to_numeric(base["Unit Price"], errors="coerce")

    # Add per-week columns
    for lbl in display_labels:
        ws = label_to_start.get(lbl)
        if not ws:
            base[lbl] = pd.NA
            continue
        sub = wk[wk["week_start"] == ws][["sku","UnitsResolved"]].rename(columns={"sku":"SKU", "UnitsResolved": lbl})
        base = base.merge(sub, on="SKU", how="left")

    # Add Sales/Notes for edit week only (far right)
    edit_start = label_to_start.get(edit_label)
    if edit_start and not wk.empty:
        sub2 = wk[wk["week_start"] == edit_start][["sku","sales_manual","notes"]].rename(columns={"sku":"SKU"})
        base = base.merge(sub2, on="SKU", how="left")
    else:
        base["sales_manual"] = pd.NA
        base["notes"] = pd.NA

    base.rename(columns={"sales_manual":"Sales", "notes":"Notes"}, inplace=True)

    # Total $ across displayed weeks (calculated, read-only)
    # Sum units across displayed labels * unit price per row
    units_sum = None
    for lbl in display_labels:
        col = pd.to_numeric(base[lbl], errors="coerce")
        units_sum = col if units_sum is None else units_sum.add(col, fill_value=0)
        # Total $ is the Δ $ between the last two displayed weeks: (last - prev) × Unit Price
    if len(display_labels) >= 2:
        prev_lbl = display_labels[-2]
        last_lbl = display_labels[-1]
        prev_u = pd.to_numeric(base[prev_lbl], errors="coerce").fillna(0)
        last_u = pd.to_numeric(base[last_lbl], errors="coerce").fillna(0)
        base["Total $ (Units x Price)"] = ((last_u - prev_u) * base["Unit Price"].fillna(0)).where(base["Unit Price"].notna(), pd.NA)
        base["Total $ (Units x Price)"] = pd.to_numeric(base["Total $ (Units x Price)"], errors="coerce").round(2)
    else:
        base["Total $ (Units x Price)"] = pd.NA
    # Δ Units between the last two displayed weeks (per SKU)
    if len(display_labels) >= 2:
        prev_lbl = display_labels[-2]
        last_lbl = display_labels[-1]
        prev_vals = pd.to_numeric(base[prev_lbl], errors="coerce").fillna(0)
        last_vals = pd.to_numeric(base[last_lbl], errors="coerce").fillna(0)
        base["Δ Units (Last - Prev)"] = last_vals - prev_vals
    else:
        base["Δ Units (Last - Prev)"] = pd.NA



    # Reorder columns: Vendor, SKU, Unit Price, week cols..., Total$, Sales, Notes
    cols = ["Vendor","SKU","Unit Price"] + display_labels + ["Total $ (Units x Price)","Sales","Notes","Δ Units (Last - Prev)"]
    return base[cols]

def save_edit_week(conn, retailer: str, week_start: date, week_end: date, edit_label: str, edited_df: pd.DataFrame):
    now = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    cur = conn.cursor()

    # The editable units are in the column named edit_label.
    for _, row in edited_df.iterrows():
        sku = str(row["SKU"]).strip()
        units_val = row.get(edit_label)
        sales_val = row.get("Sales")
        notes_val = row.get("Notes")

        # Units override stored as the edited value (can be blank to clear)
        units_override = None
        if units_val is not None and not (isinstance(units_val, float) and pd.isna(units_val)):
            try:
                units_override = float(units_val)
            except Exception:
                units_override = None

        sales_manual = None
        if sales_val is not None and not (isinstance(sales_val, float) and pd.isna(sales_val)):
            try:
                sales_manual = float(sales_val)
            except Exception:
                sales_manual = None

        notes_txt = None
        if notes_val is not None and not (isinstance(notes_val, float) and pd.isna(notes_val)) and str(notes_val).strip() != "":
            notes_txt = str(notes_val)

        cur.execute("""
            INSERT INTO weekly_results(week_start, week_end, retailer, sku,
                                      units_auto, units_override, sales_manual, notes, updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(week_start, retailer, sku) DO UPDATE SET
                week_end=excluded.week_end,
                units_override=excluded.units_override,
                sales_manual=excluded.sales_manual,
                notes=excluded.notes,
                updated_at=excluded.updated_at
        """, (
            week_start.isoformat(), week_end.isoformat(), retailer, sku,
            None, units_override, sales_manual, notes_txt, now
        ))
    conn.commit()

# -----------------------------
# UI
# -----------------------------
# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.markdown(COMPACT_TABLE_CSS, unsafe_allow_html=True)
st.markdown(SPLIT_TABLE_CSS, unsafe_allow_html=True)
st.title(APP_TITLE)


# compact grid
st.markdown(
    """
    <style>
      /* compact grid */
      div[data-testid="stDataFrame"] thead tr th, 
      div[data-testid="stDataFrame"] tbody tr td {
        padding-top: 0.15rem !important;
        padding-bottom: 0.15rem !important;
        padding-left: 0.35rem !important;
        padding-right: 0.35rem !important;
        white-space: nowrap !important;
      }
      /* allow table to size to content */
      div[data-testid="stDataFrame"] [role="grid"] {
        width: max-content !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)


# Light padding so the table fits more rows
st.markdown(
    """
    <style>
      .block-container { padding-top: 1.0rem; padding-bottom: 1.0rem; }
      section[data-testid="stSidebar"] .block-container { padding-top: 0.8rem; }
    </style>
    """,
    unsafe_allow_html=True
)

conn = get_conn()
init_db(conn)
ensure_mapping_loaded(conn, os.path.join(os.path.dirname(__file__), "Vendor-SKU Map.xlsx"))
booted = bootstrap_default_mapping(conn)
refreshed_prices = refresh_mapping_from_bundled_if_needed(conn)

# Load retailers ONCE, before any sidebar/main references
retailers = get_retailers(conn)
if not retailers:
    st.error("No retailers found. Make sure 'Vendor-SKU Map.xlsx' is in the repo (or upload one in the sidebar).")
    st.stop()

week_meta = weeks_2026()
labels = [w[2] for w in week_meta]

# -----------------------------
# Sidebar controls (retailer/weeks/upload)
# -----------------------------
with st.sidebar:
    st.header("Setup (optional)")
    st.caption("Vendor map is bundled in the repo. Upload only if you want to replace it.")
    st.write("✅ Loaded bundled mapping" if booted else "ℹ️ Using existing mapping in database")
    if refreshed_prices:
        st.success("Prices refreshed from bundled vendor map")

    map_file = st.file_uploader("Upload Vendor-SKU Map (.xlsx) (optional)", type=["xlsx"])
    if map_file is not None:
        try:
            df_map = pd.read_excel(map_file, sheet_name=0)
            upsert_mapping(conn, df_map)
            st.success("Mapping updated. Reloading retailers…")
            retailers = get_retailers(conn)
        except Exception as e:
            st.error(f"Mapping upload failed: {e}")

    st.divider()
    st.header("Report controls")

    retailer = st.selectbox("Retailer", retailers, key="retailer_sidebar")

    # Restore saved selections per retailer
    state_key = f"ui::{retailer}"
    saved = get_ui_state(conn, state_key, default={}) or {}
    saved_display = saved.get("display_weeks")
    saved_edit = saved.get("edit_week")

    default_display = labels[:9]  # partial + first 8 full weeks
    if isinstance(saved_display, list):
        default_display = [w for w in labels if w in saved_display] or default_display

    selection_mode = st.radio("Select by", ["Weeks", "Months", "Both"], index=0 if saved.get("selection_mode") is None else ["Weeks","Months","Both"].index(saved.get("selection_mode")), horizontal=True)

    saved_months = saved.get("months") if isinstance(saved.get("months"), list) else []
    month_labels = [MONTH_NAMES[m-1] for m in range(1, 13)]
    month_label_to_num = {MONTH_NAMES[m-1]: m for m in range(1, 13)}
    default_month_labels = [MONTH_NAMES[m-1] for m in saved_months if 1 <= m <= 12]
    month_sel_labels = st.multiselect("Months", month_labels, default=default_month_labels, key="months_sidebar")
    months_selected = [month_label_to_num[x] for x in month_sel_labels]

    display_weeks_manual = st.multiselect("Weeks to display", labels, default=default_display, key="display_weeks_sidebar")
    display_weeks_manual = [lbl for lbl in labels if lbl in display_weeks_manual]  # chronological

    display_weeks_by_month = weeks_for_months(week_meta, months_selected)
    display_weeks_by_month = [lbl for lbl in labels if lbl in display_weeks_by_month]  # chronological

    if selection_mode == "Weeks":
        display_weeks = display_weeks_manual
    elif selection_mode == "Months":
        display_weeks = display_weeks_by_month
    else:
        display_weeks = [lbl for lbl in labels if (lbl in set(display_weeks_manual) or lbl in set(display_weeks_by_month))]

    edit_index = labels.index(saved_edit) if (saved_edit in labels) else 0
    edit_week = st.selectbox("Week to edit", labels, index=edit_index, key="edit_week_sidebar")

    if edit_week not in display_weeks:
        display_weeks = display_weeks + [edit_week]

    # Persist selections every run
    set_ui_state(conn, state_key, {"display_weeks": display_weeks, "edit_week": edit_week, "months": months_selected, "selection_mode": selection_mode})

    st.divider()
    st.subheader("Upload units (APP workbook)")
    app_file = st.file_uploader("Weekly APP workbook (.xlsx)", type=["xlsx"], key="app_units_upload")

    if app_file is not None:
        parsed_label, _, _ = parse_week_label_from_filename(getattr(app_file, "name", ""))
        if parsed_label:
            st.caption(f"Detected week in filename: {parsed_label}")

        if st.button("Import units into Edit Week", type="primary", use_container_width=False):
            chosen_start, chosen_end, _ = next((a, b, l) for a, b, l in week_meta if l == edit_week)

            wb_up = load_workbook(app_file, data_only=True)
            imported = []
            skipped = []
            for sh in wb_up.sheetnames:
                retailer_name = normalize_retailer_sheet_name(sh)
                ws = wb_up[sh]
                units = parse_app_aggregated_sheet(ws)
                if not units:
                    skipped.append(sh)
                    continue
                set_units_auto_from_upload(conn, chosen_start, chosen_end, retailer_name, units)
                imported.append((retailer_name, len(units)))

            if imported:
                msg = ", ".join([f"{r} ({n})" for r, n in imported])
                st.success(f"Imported: {msg}")
            if skipped:
                st.caption(f"Skipped empty sheets: {', '.join(skipped)}")

# -----------------------------
# Main tabs
# -----------------------------

# Global toggles (apply across all tabs)
if "global_edit_mode" not in st.session_state:
    st.session_state["global_edit_mode"] = True
if "global_color_deltas" not in st.session_state:
    st.session_state["global_color_deltas"] = True

st.toggle("Edit mode (applies to all tabs)", key="global_edit_mode")
st.toggle("Color positive/negative deltas", key="global_color_deltas")

tab_report, tab_summary, tab_top_retailer, tab_top_vendor, tab_total_sku = st.tabs(["Report", "Summary", "Top 10 by Retailer", "Top SKU by Vendor", "Total $ per SKU"])

with tab_report:
    edit_mode = st.session_state.get('global_edit_mode', True)
    color_deltas = st.session_state.get('global_color_deltas', True)
    st.markdown(f"**Retailer:** {retailer}  |  **Edit week:** {edit_week}  |  **Weeks shown:** {', '.join(display_weeks)}")

    # Build and render table
    df = build_multiweek_df(conn, retailer, week_meta, display_weeks, edit_week)

    # Optional filter: only show items with activity
    show_only_with_units = st.checkbox("Show only items with units (or sales)", value=True)
    if show_only_with_units and not df.empty:
        week_cols = [c for c in display_weeks if c in df.columns]
        units_any = (df[week_cols].apply(pd.to_numeric, errors="coerce").fillna(0) > 0).any(axis=1) if week_cols else pd.Series([False] * len(df))
        sales_any = pd.to_numeric(df["Sales"], errors="coerce").fillna(0) != 0
        df = df[units_any | sales_any].reset_index(drop=True)

    if df.empty:
        st.info("No rows for this retailer in your mapping (or no activity for the selected weeks).")
        st.stop()

    # Keep Unit Price for calculations, but hide it from the table
    unit_price = pd.to_numeric(df["Unit Price"], errors="coerce").fillna(0)
    df = df.drop(columns=["Unit Price"])
    # Keep Sales and Notes for persistence, but hide from the table
    hidden_sales = pd.to_numeric(df["Sales"], errors="coerce") if "Sales" in df.columns else None
    df = df.drop(columns=[c for c in ["Sales", "Notes"] if c in df.columns])

    # Ensure all money columns are numeric + rounded (prevents long float tails)
    money_cols_all = [c for c in df.columns if "$" in c]
    for c in money_cols_all:
        df[c] = pd.to_numeric(df[c], errors="coerce").round(2)

    # Disable columns: keep Vendor/SKU and non-edit weeks read-only
    disabled_cols = ["Vendor", "SKU", "Total $ (Units x Price)", "Δ Units (Last - Prev)"] + [w for w in display_weeks if w != edit_week]

    if not edit_mode:
        view_df = df.copy()

        # Units table (numeric)
        view_units = view_df.copy()
        # Units table should not include any $ columns (sales are shown in the separate dollars table)
        view_units = view_units[[c for c in view_units.columns if '$' not in c]]

        # Dollars table: Vendor, SKU + $ per selected week (units * unit_price)
        view_dollars = pd.DataFrame({
            "Vendor": view_units.get("Vendor", ""),
            "SKU": view_units.get("SKU", ""),
        })
        for w in display_weeks:
            if w in view_units.columns:
                u = pd.to_numeric(view_units[w], errors="coerce").fillna(0)
                view_dollars[w] = (u * unit_price).round(2)

        # Δ $ between last two weeks
        if len(display_weeks) >= 2:
            prev_w, last_w = display_weeks[-2], display_weeks[-1]
            if prev_w in view_dollars.columns and last_w in view_dollars.columns:
                view_dollars["Δ $ (Last - Prev)"] = (
                    pd.to_numeric(view_dollars[last_w], errors="coerce").fillna(0)
                    - pd.to_numeric(view_dollars[prev_w], errors="coerce").fillna(0)
                ).round(2)
            else:
                view_dollars["Δ $ (Last - Prev)"] = pd.NA
        else:
            view_dollars["Δ $ (Last - Prev)"] = pd.NA

        # Currency formatting as strings (isolated to dollars table)
        # Append TOTAL row to both tables (based on shown rows)
        week_cols_units = [w for w in display_weeks if w in view_units.columns]
        tot_units = {w: float(pd.to_numeric(view_units[w], errors='coerce').fillna(0).sum()) for w in week_cols_units}
        if len(week_cols_units) >= 2:
            prev_w, last_w = week_cols_units[-2], week_cols_units[-1]
            delta_units_total = tot_units[last_w] - tot_units[prev_w]
        else:
            delta_units_total = pd.NA
        totals_units_row = {'Vendor': 'TOTAL', 'SKU': ''}
        totals_units_row.update(tot_units)
        if 'Δ Units (Last - Prev)' in view_units.columns:
            totals_units_row['Δ Units (Last - Prev)'] = delta_units_total
        view_units = pd.concat([view_units, pd.DataFrame([totals_units_row])], ignore_index=True)
        
        week_cols_dollars = [w for w in display_weeks if w in view_dollars.columns]
        tot_dollars = {w: float(pd.to_numeric(view_dollars[w], errors='coerce').fillna(0).sum()) for w in week_cols_dollars}
        if len(week_cols_dollars) >= 2:
            prev_w, last_w = week_cols_dollars[-2], week_cols_dollars[-1]
            delta_dollars_total = tot_dollars[last_w] - tot_dollars[prev_w]
        else:
            delta_dollars_total = pd.NA
        totals_dollars_row = {'Vendor': 'TOTAL', 'SKU': ''}
        totals_dollars_row.update(tot_dollars)
        totals_dollars_row['Δ $ (Last - Prev)'] = delta_dollars_total
        view_dollars = pd.concat([view_dollars, pd.DataFrame([totals_dollars_row])], ignore_index=True)
        
        for c in [w for w in display_weeks if w in view_dollars.columns] + ["Δ $ (Last - Prev)"]:
            if c in view_dollars.columns:
                view_dollars[c] = pd.to_numeric(view_dollars[c], errors="coerce").round(2).apply(fmt_currency_str)

        def _color_pos_neg(val):
            try:
                v = float(str(val).replace('(','').replace(')','').replace('$','').replace(',',''))
            except Exception:
                return ""
            if v > 0:
                return "color: #1f8b4c; font-weight: 600;"
            if v < 0:
                return "color: #c92a2a; font-weight: 600;"
            return ""

        # Render two tables side-by-side
        left_col, right_col = st.columns([1, 1], gap="small")

        with left_col:
            styled_units = view_units.style
            if color_deltas and "Δ Units (Last - Prev)" in view_units.columns:
                styled_units = styled_units.applymap(_color_pos_neg, subset=["Δ Units (Last - Prev)"])
            st.dataframe(
                styled_units,
                use_container_width=False,
                height=900,
                column_config={
                    "Vendor": st.column_config.TextColumn(width="small"),
                    "SKU": st.column_config.TextColumn(width="small"),
                    **{w: st.column_config.NumberColumn(format="%.0f", width="small") for w in display_weeks if w in view_units.columns},
                    "Δ Units (Last - Prev)": st.column_config.NumberColumn(format="%.0f", width="small"),
                },
            )

        with right_col:
            styled_dollars = view_dollars.style
            if color_deltas and "Δ $ (Last - Prev)" in view_dollars.columns:
                styled_dollars = styled_dollars.applymap(_color_pos_neg, subset=["Δ $ (Last - Prev)"])
            st.dataframe(
                styled_dollars,
                use_container_width=False,
                height=900,
                hide_index=True,
                column_config={
                    "Vendor": st.column_config.TextColumn(width="small"),
                    "SKU": st.column_config.TextColumn(width="small"),
                    **{w: st.column_config.TextColumn(width="small") for w in display_weeks if w in view_dollars.columns},
                    "Δ $ (Last - Prev)": st.column_config.TextColumn(width="small"),
                },
            )

        edited = view_df

    else:
        # Editor: Units editable (only Edit Week). Dollars computed from the edited units.
        left_col, right_col = st.columns([1, 1], gap="small")

        with left_col:
            df_editor_main = df.copy()
            keep_cols = ["Vendor", "SKU"] + [w for w in display_weeks if w in df_editor_main.columns]
            if "Δ Units (Last - Prev)" in df_editor_main.columns:
                keep_cols += ["Δ Units (Last - Prev)"]
            df_editor_main = df_editor_main[keep_cols].copy()

            edited = st.data_editor(
                df_editor_main,
                height=900,
                use_container_width=False,
                hide_index=True,
                disabled=disabled_cols,
                column_config={
                    "Vendor": st.column_config.TextColumn(width="small"),
                    "SKU": st.column_config.TextColumn(width="small"),
                    **{w: st.column_config.NumberColumn(format="%.0f", width="small") for w in display_weeks if w in df_editor_main.columns},
                    "Δ Units (Last - Prev)": st.column_config.NumberColumn(format="%.0f", width="small"),
                },
            )

        with right_col:
            dollars = pd.DataFrame({
                "Vendor": edited.get("Vendor", ""),
                "SKU": edited.get("SKU", ""),
            })
            for w in display_weeks:
                if w in edited.columns:
                    u = pd.to_numeric(edited[w], errors="coerce").fillna(0)
                    dollars[w] = (u * unit_price).round(2)

            if len(display_weeks) >= 2:
                prev_w, last_w = display_weeks[-2], display_weeks[-1]
                if prev_w in dollars.columns and last_w in dollars.columns:
                    dollars["Δ $ (Last - Prev)"] = (
                        pd.to_numeric(dollars[last_w], errors="coerce").fillna(0)
                        - pd.to_numeric(dollars[prev_w], errors="coerce").fillna(0)
                    ).round(2)
                else:
                    dollars["Δ $ (Last - Prev)"] = pd.NA
            else:
                dollars["Δ $ (Last - Prev)"] = pd.NA

            # Append TOTAL row (dollars table) based on edited rows
            week_cols_dollars = [w for w in display_weeks if w in dollars.columns]
            tot_dollars = {w: float(pd.to_numeric(dollars[w], errors='coerce').fillna(0).sum()) for w in week_cols_dollars}
            if len(week_cols_dollars) >= 2:
                prev_w, last_w = week_cols_dollars[-2], week_cols_dollars[-1]
                delta_dollars_total = tot_dollars[last_w] - tot_dollars[prev_w]
            else:
                delta_dollars_total = pd.NA
            totals_row = {'Vendor': 'TOTAL', 'SKU': ''}
            totals_row.update(tot_dollars)
            totals_row['Δ $ (Last - Prev)'] = delta_dollars_total
            dollars = pd.concat([dollars, pd.DataFrame([totals_row])], ignore_index=True)
            
            for c in [w for w in display_weeks if w in dollars.columns] + ["Δ $ (Last - Prev)"]:
                if c in dollars.columns:
                    dollars[c] = pd.to_numeric(dollars[c], errors="coerce").round(2).apply(fmt_currency_str)

            st.dataframe(
                dollars,
                use_container_width=False,
                height=900,
                hide_index=True,
                column_config={
                    "Vendor": st.column_config.TextColumn(width="small"),
                    "SKU": st.column_config.TextColumn(width="small"),
                    **{w: st.column_config.TextColumn(width="small") for w in display_weeks if w in dollars.columns},
                    "Δ $ (Last - Prev)": st.column_config.TextColumn(width="small"),
                },
            )

    c1, c2 = st.columns([1, 3])

    with c1:
        if st.button("Save edits", type="primary"):
            start, end, _ = next((a, b, l) for a, b, l in week_meta if l == edit_week)
            save_edit_week(conn, retailer, start, end, edit_week, edited)
            st.success("Saved.")
    with c2:
        st.caption("Only the selected Edit Week column is editable. Far-right column shows Δ Units (last selected week minus the previous week).")

with tab_summary:
    st.subheader("Summary by retailer and week")
    st.caption("Shows total units and total $ for each retailer for the weeks you selected in the sidebar. Totals match the Report tab totals logic, but include ALL SKUs.")

    label_to_start = {lbl: start.isoformat() for lbl, start in [(l, s) for s, _, l in week_meta]}
    # keep only selected labels that exist
    selected_labels = [lbl for lbl in display_weeks if lbl in label_to_start]
    selected_starts = [label_to_start[lbl] for lbl in selected_labels]

    if not selected_labels:
        st.info("Select at least one week in the sidebar to build the summary.")
    else:
        mapping_all = pd.read_sql_query(
            """
            SELECT retailer, sku, unit_price
            FROM sku_mapping
            WHERE active = 1
            """,
            conn
        )
        mapping_all["unit_price"] = pd.to_numeric(mapping_all["unit_price"], errors="coerce").fillna(0)

        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f"""
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            """,
            conn,
            params=selected_starts
        )

        if wk.empty:
            st.info("No unit data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            dfm = wk.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Dollars"] = dfm["Units"] * dfm["unit_price"]

            agg = dfm.groupby(["retailer", "week_start"], as_index=False).agg(
                total_units=("Units", "sum"),
                total_dollars=("Dollars", "sum"),
            )

            all_retailers = sorted(dfm["retailer"].unique().tolist())
            out = pd.DataFrame({"Retailer": all_retailers}).set_index("Retailer")

            for lbl in selected_labels:
                ws = label_to_start[lbl]
                sub = agg[agg["week_start"] == ws].set_index("retailer")
                out[f"{lbl} Units"] = sub["total_units"]
                out[f"{lbl} $"] = sub["total_dollars"]

            out = out.fillna(0)

            col_cfg = {}
            for lbl in selected_labels:
                col_cfg[f"{lbl} Units"] = st.column_config.NumberColumn(format="%.0f", width="small")
                col_cfg[f"{lbl} $"] = st.column_config.NumberColumn(format="$%,.2f", width="small")

            out_display = out.copy()
            for c in out_display.columns:
                if c.endswith(' $'):
                    out_display[c] = pd.to_numeric(out_display[c], errors='coerce').round(2).apply(fmt_currency_str)
            st.dataframe(out_display, use_container_width=False, column_config=col_cfg)

            st.divider()
            st.subheader("Totals across selected weeks")

            unit_cols = [c for c in out.columns if c.endswith(" Units")]
            dol_cols = [c for c in out.columns if c.endswith(" $")]

            totals = pd.DataFrame({
                "Total Units (Selected Weeks)": out[unit_cols].sum(axis=1),
                "Total $ (Selected Weeks)": out[dol_cols].sum(axis=1),
            }).sort_values("Total $ (Selected Weeks)", ascending=False)

            totals_display = totals.copy()
            if 'Total $ (Selected Weeks)' in totals_display.columns:
                totals_display['Total $ (Selected Weeks)'] = pd.to_numeric(totals_display['Total $ (Selected Weeks)'], errors='coerce').round(2).apply(fmt_currency_str)
            st.dataframe(
                totals_display,
                use_container_width=False,
                column_config={
                    "Total Units (Selected Weeks)": st.column_config.NumberColumn(format="%.0f", width="small"),
                    "Total $ (Selected Weeks)": st.column_config.NumberColumn(format="$%,.2f", width="small"),
                }
            )
with tab_top_retailer:
    st.subheader("Top 10 items per retailer (by Units)")
    st.caption("Uses the weeks selected in 'Weeks to display'. Totals are summed across those weeks.")

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    selected_starts = [label_to_start[lbl] for lbl in display_weeks if lbl in label_to_start]

    if not selected_starts:
        st.info("Select at least one week to compute top sellers.")
    else:
        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f"""
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            """,
            conn,
            params=selected_starts
        )

        if wk.empty:
            st.info("No unit data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            agg = wk.groupby(["retailer", "sku"], as_index=False)["Units"].sum()

            mapping_all = pd.read_sql_query(
                """
                SELECT retailer, vendor, sku
                FROM sku_mapping
                WHERE active = 1
                """,
                conn
            )
            dfm = agg.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["vendor"] = dfm["vendor"].fillna("Unknown")

            for ret in sorted(dfm["retailer"].unique().tolist()):
                st.markdown(f"### {ret}")
                sub = dfm[dfm["retailer"] == ret].copy()
                top = sub.sort_values("Units", ascending=False).head(10)

                out = top.rename(columns={
                    "vendor": "Vendor",
                    "sku": "SKU",
                    "Units": "Total Units (Selected Weeks)"
                })[["SKU", "Vendor", "Total Units (Selected Weeks)"]]
                st.dataframe(out, use_container_width=False, column_config={"Total Units (Selected Weeks)": st.column_config.NumberColumn(format="%.0f", width="small"),"SKU": st.column_config.TextColumn(width="small"),"Vendor": st.column_config.TextColumn(width="small"),})

with tab_top_vendor:
    st.subheader("Top SKU per vendor, per retailer (by Units)")
    st.caption("Uses the weeks selected in 'Weeks to display'. Totals are summed across those weeks.")

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    selected_starts = [label_to_start[lbl] for lbl in display_weeks if lbl in label_to_start]

    if not selected_starts:
        st.info("Select at least one week to compute top sellers.")
    else:
        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f"""
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            """,
            conn,
            params=selected_starts
        )

        if wk.empty:
            st.info("No unit data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            agg = wk.groupby(["retailer", "sku"], as_index=False)["Units"].sum()

            mapping_all = pd.read_sql_query(
                """
                SELECT retailer, vendor, sku
                FROM sku_mapping
                WHERE active = 1
                """,
                conn
            )
            dfm = agg.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["vendor"] = dfm["vendor"].fillna("Unknown")

            for ret in sorted(dfm["retailer"].unique().tolist()):
                st.markdown(f"### {ret}")
                subr = dfm[dfm["retailer"] == ret].copy()

                for vend in sorted(subr["vendor"].unique().tolist()):
                    subv = subr[subr["vendor"] == vend].copy()
                    if subv.empty:
                        continue

                    top = subv.sort_values("Units", ascending=False).head(1)

                    out = top.rename(columns={
                        "sku": "SKU",
                        "Units": "Total Units (Selected Weeks)"
                    })[["SKU", "Total Units (Selected Weeks)"]]

                    st.write(f"**{vend}**")
                    st.dataframe(out, use_container_width=False, column_config={"Total Units (Selected Weeks)": st.column_config.NumberColumn(format="%.0f", width="small"),"SKU": st.column_config.TextColumn(width="small"),"Vendor": st.column_config.TextColumn(width="small"),})
with tab_total_sku:
    st.subheader("Total $ per SKU (selected weeks)")
    st.caption("Totals are summed across the weeks you selected in the sidebar. This uses Unit Price from the vendor map.")

    label_to_start = {lbl: start.isoformat() for start, _, lbl in week_meta}
    selected_starts = [label_to_start[lbl] for lbl in display_weeks if lbl in label_to_start]

    if not selected_starts:
        st.info("Select at least one week to compute totals.")
    else:
        placeholders = ",".join(["?"] * len(selected_starts))
        wk = pd.read_sql_query(
            f"""
            SELECT week_start, retailer, sku, units_auto, units_override
            FROM weekly_results
            WHERE week_start IN ({placeholders})
            """,
            conn,
            params=selected_starts
        )

        if wk.empty:
            st.info("No unit data found for the selected weeks yet.")
        else:
            wk["Units"] = wk["units_override"].where(wk["units_override"].notna(), wk["units_auto"])
            wk["Units"] = pd.to_numeric(wk["Units"], errors="coerce").fillna(0)

            agg = wk.groupby(["retailer", "sku"], as_index=False)["Units"].sum()

            mapping_all = pd.read_sql_query(
                """
                SELECT retailer, vendor, sku, unit_price
                FROM sku_mapping
                WHERE active = 1
                """,
                conn
            )
            dfm = agg.merge(mapping_all, on=["retailer", "sku"], how="left")
            dfm["vendor"] = dfm["vendor"].fillna("Unknown")
            dfm["unit_price"] = pd.to_numeric(dfm["unit_price"], errors="coerce").fillna(0)
            dfm["Total $"] = (dfm["Units"] * dfm["unit_price"]).round(2)

            for ret in sorted(dfm["retailer"].unique().tolist()):
                st.markdown(f"### {ret}")
                sub = dfm[dfm["retailer"] == ret].copy()
                sub = sub.sort_values("Total $", ascending=False)

                out = sub.rename(columns={
                    "sku": "SKU",
                    "vendor": "Vendor",
                    "Units": "Total Units",
                    "unit_price": "Unit Price",
                    "Total $": "Total $"
                })[["SKU", "Vendor", "Total Units", "Unit Price", "Total $"]]

                out_display = out.copy()
                if 'Unit Price' in out_display.columns:
                    out_display['Unit Price'] = pd.to_numeric(out_display['Unit Price'], errors='coerce').round(2).apply(fmt_currency_str)
                if 'Total $' in out_display.columns:
                    out_display['Total $'] = pd.to_numeric(out_display['Total $'], errors='coerce').round(2).apply(fmt_currency_str)
                st.dataframe(
                    out_display,
                    use_container_width=False,
                    column_config={
                        "Unit Price": st.column_config.NumberColumn(format="$%,.2f", width="small"),
                        "Total $": st.column_config.NumberColumn(format="$%,.2f", width="small"),
                        "Total Units": st.column_config.NumberColumn(format="%.0f", width="small"),
                    }
                )